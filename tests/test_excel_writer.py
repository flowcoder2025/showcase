import pandas as pd
import pytest
from openpyxl import load_workbook

from core.excel import writer


def test_write_styled_report_creates_file(tmp_path):
    df = pd.DataFrame(
        {"vendor": ["A", "B"], "2026-01": [100, 200], "2026-02": [300, 400]}
    ).set_index("vendor")
    out = tmp_path / "report.xlsx"
    writer.write_styled_report(df, out, title="거래처 매출")
    assert out.exists()
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "거래처 매출"
    # 차트 1개 이상 존재 (broken chart 회피)
    assert len(ws._charts) >= 1, "차트가 생성되지 않음"


def test_write_styled_report_rejects_multiindex(tmp_path):
    """Writer requires single-level index — reject MultiIndex up front."""
    multi_df = pd.DataFrame({
        "amount": [100, 200, 300, 400],
    }, index=pd.MultiIndex.from_tuples([
        ("A", "Seoul"), ("A", "Busan"), ("B", "Seoul"), ("B", "Busan"),
    ], names=["vendor", "region"]))
    out = tmp_path / "multi.xlsx"
    with pytest.raises(ValueError, match="single-level index"):
        writer.write_styled_report(multi_df, out)
