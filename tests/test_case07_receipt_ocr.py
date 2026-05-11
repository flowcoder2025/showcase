"""T11: case07 — 영수증 일괄 OCR → 경비 정리 엑셀 (T38 ScenarioResult).

``core.ocr.receipt.extract`` mock 기반 contract 검증. 실제 Ollama 호출 없음.
output_dir 기반 (T38) — 결과 xlsx 파일은 ``result["output_files"][0]`` 로 접근.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pytest
from flowcoder_office_tools.ocr import receipt
from flowcoder_office_tools.ocr.receipt import ReceiptData
from PIL import Image

from cases.case07_ocr_receipt_to_excel import scenario


def _make_blank_png(path: Path) -> None:
    Image.new("RGB", (10, 10), "white").save(path)


def _mock_receipt(
    monkeypatch: pytest.MonkeyPatch,
    response: ReceiptData | None = None,
    fail_filenames: tuple[str, ...] = (),
) -> list[Path]:
    calls: list[Path] = []
    default: ReceiptData = response or ReceiptData(
        merchant="스타벅스 강남점",
        amount=5500,
        date="2026-04-15",
        items=[{"name": "아메리카노", "qty": 1, "price": 5500}],
        raw_text="스타벅스 5500",
    )

    def _fake(image_path: Path | str) -> ReceiptData:
        p = Path(image_path)
        calls.append(p)
        if p.name in fail_filenames:
            raise ValueError(f"mock OCR failure for {p.name}")
        return default

    monkeypatch.setattr(receipt, "extract", _fake)
    return calls


def _output_xlsx(result: dict) -> Path:
    out_path = result["output_files"][0]
    assert out_path.name == "expense_report.xlsx"
    return out_path


# -- 1. 기본 처리 ----------------------------------------------------------


def test_run_processes_all_images_in_input_dir(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    input_dir = tmp_path / "in"
    input_dir.mkdir()
    for i in range(5):
        _make_blank_png(input_dir / f"r{i:03d}.png")

    _mock_receipt(monkeypatch)
    out_dir = tmp_path / "out"

    result = scenario.run(input_dir=input_dir, output_dir=out_dir)
    out_path = _output_xlsx(result)

    assert result["metrics"]["processed"] == 5
    assert result["metrics"]["errors"] == 0
    assert out_path.exists()

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    assert ws is not None
    assert ws.max_row == 6


# -- 2. personas fallback --------------------------------------------------


def test_run_uses_personas_fallback_when_input_empty(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    seed_dir = tmp_path / "seeds"
    seed_dir.mkdir()
    for i in range(3):
        _make_blank_png(seed_dir / f"r{i:03d}.png")
    (seed_dir / "_ground_truth.json").write_text("[]", encoding="utf-8")

    monkeypatch.setattr(scenario, "_DEFAULT_IN", seed_dir)
    _mock_receipt(monkeypatch)

    out_dir = tmp_path / "out"
    result = scenario.run(input_dir=None, output_dir=out_dir)

    assert result["metrics"]["processed"] == 3
    assert result["metrics"]["errors"] == 0


# -- 3. per-image 실패 격리 -------------------------------------------------


def test_run_continues_after_per_image_failure(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    input_dir = tmp_path / "in"
    input_dir.mkdir()
    for i in range(4):
        _make_blank_png(input_dir / f"r{i:03d}.png")

    _mock_receipt(monkeypatch, fail_filenames=("r001.png", "r003.png"))
    out_dir = tmp_path / "out"

    result = scenario.run(input_dir=input_dir, output_dir=out_dir)
    out_path = _output_xlsx(result)

    assert result["metrics"]["processed"] == 2
    assert result["metrics"]["errors"] == 2

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    assert ws is not None
    assert ws.max_row == 3


# -- 4. safe-fallback 데이터 ------------------------------------------------


def test_run_safe_mode_returns_safe_fallback_data(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    input_dir = tmp_path / "in"
    input_dir.mkdir()
    for i in range(4):
        _make_blank_png(input_dir / f"r{i:03d}.png")

    safe_data: ReceiptData = ReceiptData(
        merchant="[SAFE-FALLBACK]",
        amount=0,
        date="2026-01-01",
        items=[],
        raw_text="safe_dummy: abc123",
    )
    _mock_receipt(monkeypatch, response=safe_data)
    out_dir = tmp_path / "out"

    result = scenario.run(input_dir=input_dir, output_dir=out_dir)
    out_path = _output_xlsx(result)

    assert result["metrics"]["processed"] == 4
    assert result["metrics"]["errors"] == 0

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    assert ws is not None
    merchants = [ws.cell(row=r, column=2).value for r in range(2, 6)]
    assert all(m == "[SAFE-FALLBACK]" for m in merchants)


# -- 5. underscore prefix 스킵 ----------------------------------------------


def test_run_skips_underscore_prefix_files(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    input_dir = tmp_path / "in"
    input_dir.mkdir()
    _make_blank_png(input_dir / "r001.png")
    _make_blank_png(input_dir / "r002.png")
    _make_blank_png(input_dir / "_temp.png")
    (input_dir / "_ground_truth.json").write_text("[]", encoding="utf-8")

    _mock_receipt(monkeypatch)
    out_dir = tmp_path / "out"

    result = scenario.run(input_dir=input_dir, output_dir=out_dir)

    assert result["metrics"]["processed"] == 2
    assert result["metrics"]["errors"] == 0


# -- 6. xlsx 컬럼 헤더 ------------------------------------------------------


def test_run_output_xlsx_columns(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    input_dir = tmp_path / "in"
    input_dir.mkdir()
    _make_blank_png(input_dir / "r001.png")

    _mock_receipt(monkeypatch)
    out_dir = tmp_path / "out"
    result = scenario.run(input_dir=input_dir, output_dir=out_dir)
    out_path = _output_xlsx(result)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    assert ws is not None
    headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
    assert headers == ["거래일", "가맹점", "카테고리", "결제수단", "금액"]


# -- 7. 카테고리 매핑 ------------------------------------------------------


def test_guess_category_known_prefixes() -> None:
    assert scenario._guess_category("스타벅스 강남점") == "커피"
    assert scenario._guess_category("이디야 역삼") == "커피"
    assert scenario._guess_category("롯데마트 송파") == "장보기"
    assert scenario._guess_category("이마트 트레이더스") == "장보기"
    assert scenario._guess_category("GS25 서초") == "편의점"
    assert scenario._guess_category("CU 청담") == "편의점"
    assert scenario._guess_category("세븐일레븐 합정") == "편의점"
    assert scenario._guess_category("맥도날드 강남대로") == "식사"
    assert scenario._guess_category("알 수 없는 가게") == "기타"


# -- 8. 빈 디렉토리 ---------------------------------------------------------


def test_run_zero_images(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    input_dir = tmp_path / "in"
    input_dir.mkdir()

    empty_seed = tmp_path / "empty_seed"
    empty_seed.mkdir()
    monkeypatch.setattr(scenario, "_DEFAULT_IN", empty_seed)

    _mock_receipt(monkeypatch)
    out_dir = tmp_path / "out"

    result = scenario.run(input_dir=input_dir, output_dir=out_dir)
    out_path = _output_xlsx(result)

    assert result["metrics"]["processed"] == 0
    assert result["metrics"]["errors"] == 0
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    assert ws is not None
    assert ws.max_row == 1


# -- 9. 이미지 확장자 필터 -------------------------------------------------


def test_run_filters_only_image_extensions(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    input_dir = tmp_path / "in"
    input_dir.mkdir()
    _make_blank_png(input_dir / "r001.png")
    _make_blank_png(input_dir / "r002.jpg")
    _make_blank_png(input_dir / "r003.jpeg")
    (input_dir / "notes.txt").write_text("ignore me", encoding="utf-8")
    (input_dir / "doc.pdf").write_bytes(b"%PDF-1.4 fake")

    _mock_receipt(monkeypatch)
    out_dir = tmp_path / "out"
    result = scenario.run(input_dir=input_dir, output_dir=out_dir)

    assert result["metrics"]["processed"] == 3
    assert result["metrics"]["errors"] == 0


# -- 10. output 디렉토리 자동 생성 -----------------------------------------


def test_run_creates_output_dir_if_missing(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    input_dir = tmp_path / "in"
    input_dir.mkdir()
    _make_blank_png(input_dir / "r001.png")

    _mock_receipt(monkeypatch)
    out_dir = tmp_path / "deep" / "nested" / "out"
    assert not out_dir.exists()

    result = scenario.run(input_dir=input_dir, output_dir=out_dir)
    out_path = _output_xlsx(result)
    assert out_path.exists()


# -- 11. extras["receipts"] 구조 -------------------------------------------


def test_run_extras_receipts_structure(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    input_dir = tmp_path / "in"
    input_dir.mkdir()
    _make_blank_png(input_dir / "r001.png")
    _make_blank_png(input_dir / "r002.png")

    _mock_receipt(monkeypatch)
    out_dir = tmp_path / "out"
    result = scenario.run(input_dir=input_dir, output_dir=out_dir)

    receipts: list[dict[str, Any]] = result["extras"]["receipts"]
    assert len(receipts) == 2
    for row in receipts:
        assert "filename" in row
        assert "merchant" in row
        assert "amount" in row


# -- 12. 금액은 정수 ---------------------------------------------------------


def test_run_xlsx_amount_is_integer(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    input_dir = tmp_path / "in"
    input_dir.mkdir()
    _make_blank_png(input_dir / "r001.png")

    _mock_receipt(monkeypatch)
    out_dir = tmp_path / "out"
    result = scenario.run(input_dir=input_dir, output_dir=out_dir)
    out_path = _output_xlsx(result)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    assert ws is not None
    amount_cell = ws.cell(row=2, column=5).value
    assert isinstance(amount_cell, int), f"expected int, got {type(amount_cell).__name__}"
    assert amount_cell == 5500


# -- 13. T11.5: 결제수단 추출 -----------------------------------------------


def test_run_extracts_payment_from_raw_text(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    input_dir = tmp_path / "in"
    input_dir.mkdir()
    _make_blank_png(input_dir / "r001.png")

    samsung_data = ReceiptData(
        merchant="스타벅스 강남점",
        amount=5500,
        date="2026-04-15",
        items=[],
        raw_text="스타벅스 5500 삼성페이 결제 완료",
    )
    _mock_receipt(monkeypatch, response=samsung_data)
    out_dir = tmp_path / "out"
    result = scenario.run(input_dir=input_dir, output_dir=out_dir)
    out_path = _output_xlsx(result)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    assert ws is not None
    payment_cell = ws.cell(row=2, column=4).value
    assert payment_cell == "삼성페이"


def test_run_payment_unknown_returns_empty(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    input_dir = tmp_path / "in"
    input_dir.mkdir()
    _make_blank_png(input_dir / "r001.png")

    no_payment = ReceiptData(
        merchant="스타벅스",
        amount=5500,
        date="2026-04-15",
        items=[],
        raw_text="스타벅스 5500 영수증 발행",
    )
    _mock_receipt(monkeypatch, response=no_payment)
    out_dir = tmp_path / "out"
    result = scenario.run(input_dir=input_dir, output_dir=out_dir)
    out_path = _output_xlsx(result)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    assert ws is not None
    payment_cell = ws.cell(row=2, column=4).value
    assert payment_cell in (None, "")


def test_guess_payment_credit_card() -> None:
    assert scenario._guess_payment("VISA 카드결제") == "신용카드"
    assert scenario._guess_payment("신용카드 일시불") == "신용카드"
    assert scenario._guess_payment("MASTER 결제") == "신용카드"
    assert scenario._guess_payment("체크카드 승인") == "신용카드"


def test_guess_payment_naver() -> None:
    assert scenario._guess_payment("NPAY 결제") == "네이버페이"
    assert scenario._guess_payment("네이버페이로 결제했습니다") == "네이버페이"
    assert scenario._guess_payment("N페이 사용") == "네이버페이"


def test_guess_payment_kakao() -> None:
    assert scenario._guess_payment("카카오페이 머니") == "카카오페이"
    assert scenario._guess_payment("KPAY 승인") == "카카오페이"
    assert scenario._guess_payment("K페이 결제") == "카카오페이"


def test_guess_payment_cash() -> None:
    assert scenario._guess_payment("현금영수증 발행") == "현금"
    assert scenario._guess_payment("CASH PAID") == "현금"


def test_guess_payment_no_match_returns_empty() -> None:
    assert scenario._guess_payment("그냥 텍스트") == ""
    assert scenario._guess_payment("") == ""


def test_guess_payment_priority_order() -> None:
    assert scenario._guess_payment("VISA 카드결제 / 카카오페이도 가능") == "신용카드"


# -- 14. T11.5: per-image timer --------------------------------------------


def test_run_extras_per_image_ms(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    input_dir = tmp_path / "in"
    input_dir.mkdir()
    for i in range(5):
        _make_blank_png(input_dir / f"r{i:03d}.png")

    _mock_receipt(monkeypatch)
    out_dir = tmp_path / "out"
    result = scenario.run(input_dir=input_dir, output_dir=out_dir)

    per_image: list[dict[str, Any]] = result["extras"]["per_image_ms"]
    assert len(per_image) == 5
    for entry in per_image:
        assert "filename" in entry
        assert "elapsed_ms" in entry
        assert isinstance(entry["elapsed_ms"], float)
        assert entry["elapsed_ms"] >= 0.0


def test_run_per_image_ms_includes_error_flag(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    input_dir = tmp_path / "in"
    input_dir.mkdir()
    for i in range(4):
        _make_blank_png(input_dir / f"r{i:03d}.png")

    _mock_receipt(monkeypatch, fail_filenames=("r001.png", "r003.png"))
    out_dir = tmp_path / "out"
    result = scenario.run(input_dir=input_dir, output_dir=out_dir)

    per_image: list[dict[str, Any]] = result["extras"]["per_image_ms"]
    assert len(per_image) == 4
    by_name = {e["filename"]: e for e in per_image}
    assert by_name["r001.png"].get("error") is True
    assert by_name["r003.png"].get("error") is True
    assert "error" not in by_name["r000.png"] or by_name["r000.png"].get("error") is False
    assert "error" not in by_name["r002.png"] or by_name["r002.png"].get("error") is False
