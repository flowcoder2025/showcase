from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook


@pytest.fixture
def case_input(tmp_path: Path) -> Path:
    df = pd.DataFrame(
        [
            {"거래처명": "A", "거래일": "2026-01-15", "금액": 100},
            {"거래처명": "A", "거래일": "2026-02-10", "금액": 200},
            {"거래처명": "B", "거래일": "2026-01-20", "금액": 300},
        ]
    )
    p = tmp_path / "input"
    p.mkdir()
    df.to_excel(p / "data.xlsx", index=False)
    return p


def test_case01_run_produces_styled_report(case_input: Path, tmp_path: Path) -> None:
    from cases.case01_excel_vendor_report import scenario

    out = tmp_path / "output" / "report.xlsx"
    scenario.run(input_dir=case_input, output_path=out)
    assert out.exists()

    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None
    # 제목 셀
    assert ws.cell(row=1, column=1).value
