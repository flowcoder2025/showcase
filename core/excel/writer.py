"""스타일 입힌 보고서 출력 — openpyxl + 차트."""
from pathlib import Path
from typing import cast

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


def write_styled_report(df: pd.DataFrame, output_path: Path, *, title: str = "보고서") -> None:
    """피벗 결과를 헤더 스타일 + 차트와 함께 저장."""
    assert df.index.nlevels == 1, (
        "writer requires single-level index; got "
        f"{df.index.nlevels}-level MultiIndex"
    )
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_active = wb.active
    assert isinstance(ws_active, Worksheet), "Workbook() always has an active Worksheet"
    # cast past openpyxl-stubs _WorksheetOrChartsheetLike multi-inheritance quirk
    ws = cast(Worksheet, ws_active)
    ws.title = "report"

    ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)

    # 헤더
    cols = ["vendor"] + list(df.columns)
    for c_idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=3, column=c_idx, value=str(col))
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    # 데이터
    for r_idx, (vendor, row) in enumerate(df.iterrows(), start=4):
        ws.cell(row=r_idx, column=1, value=str(vendor))
        for c_idx, val in enumerate(row, start=2):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # 차트
    if len(df.columns) >= 1 and len(df.index) >= 1:
        chart = BarChart()
        chart.title = title
        data = Reference(
            ws, min_col=2, min_row=3, max_col=1 + len(df.columns), max_row=3 + len(df.index)
        )
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(df.index))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "A" + str(5 + len(df.index)))

    wb.save(output_path)
